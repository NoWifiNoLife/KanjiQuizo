# -*- coding: utf-8 -*-
"""
Created on Thu Aug  2 08:10:55 2018
@author: MICHAEL
"""

import openpyxl, random, datetime
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#from tkinter import *
from tkinter import *

class MegaExcel(object):
    #general excel things that all will need
    def __init__ (self, name_of_workbook):
        self.name_of_workbook = name_of_workbook
        workbook = load_workbook(name_of_workbook)
        self.workbook = workbook
    
    def check(something):
        while something.lower() not in 'yn':
            something = input('please enter either y or n: ')
        if something.lower() == 'y':
            return True
        return False
    
    def getSheet(self, sheet_number):
        """assumes workbook is preloaded"""
        sheet_name = self.getWB().get_sheet_names()[sheet_number]
        sheet = self.getWB().get_sheet_by_name(sheet_name)
        return sheet

    def getWBName(self):
        return self.name_of_workbook
    
    def getWB(self):
        return self.workbook
    
"""test check"""
#print(MegaExcel.check('h')) ###works!###

def makingtheBase(low, high, history_log = []):
    try:
        txtlog = open('KanjiQuizLog.txt', 'a')
        wb = MegaExcel('HeisigKanji.xlsx')
        main = wb.getSheet(0)
        correct_incorrect_log = {}
        play = input('Do you want to continue? (y/n):')
        while MegaExcel.check(play):
            random_num = random.randint(low, high)
            while random_num in history_log:
                random_num = random.randint(low, high)
            history_log.append(random_num)
            print('What is ' + main['A%s' % random_num].value + ' ?')
            correct_incorrect = input('Correct? (y/n): ')
            if correct_incorrect.lower() == 'y':
                correct_incorrect_log[random_num] = correct_incorrect_log.get(random_num, [0,0])[0] + 1
            elif correct_incorrect.lower() == 'n':
                correct_incorrect_log[random_num] = correct_incorrect_log.get(random_num, [0,0])[1] + 1
            show_answer = input('Would you like to see the answer? (y/n):')
            if MegaExcel.check(show_answer):
                print('What is ' + main['B%s' % random_num].value + ' ?')
            play = input('Do you want to continue? (y/n):')
        txtlog.write('\n' + datetime.datetime.today().strftime('%Y-%m-%d') + ' ' \
                     + datetime.datetime.today().strftime('%T') + ' %s'% history_log)
        txtlog.close()
        return correct_incorrect_log, history_log
    except Exception:
        txtlog.write('\n' + datetime.datetime.today().strftime('%Y-%m-%d') + \
                     ' ' + datetime.datetime.today().strftime('%T') + 'Error occured. recording... %s'% history_log)
        txtlog.close()
        return history_log, 'error'
    
wb = MegaExcel('HeisigKanji.xlsx')
main = wb.getSheet(0)
"""Making the GUI"""

#        random.randint()
#        text = main['A%s' % num].value, font ='Times 100'


history_log = [] #initialize with 0 so all index from len() are correct
current = 0
low, high = 250, 550

class Quiz(object):
    def __init__ (self):
        pass
    def showAnswer(event):
        global current
        global history_log
        print(current, history_log)
        n = history_log[current]
        answer =  main['B%s' %  n].value
        theLabel1['text'] = answer
        
    def prevOne(event):
        global current
        current -= 1
        n = history_log[current]
        theLabel1['text'] = main['A%s' % n].value
        
    def nextOne(event):
        global current
        random_num = random.randint(low, high)
        while random_num in history_log:
            random_num = random.randint(low,high)
        history_log.append(random_num)
        theLabel1['image'] = None
        theLabel1['text'] = main['A%s' %  random_num].value
        current = len(history_log) - 1 #This part if  global current is not called it will create a new current and fuck shit 
    
    def SadGuts(event):
        img = PhotoImage(file="SadGuts.png")
        theLabel1['image'] =  img
        theLabel1.image_names = img #I don't know why this works but its a reference thing http://effbot.org/pyfaq/why-do-my-tkinter-images-not-appear.htm
        theLabel1['text'] = 'Nope'
            
    def runtk(self):
        try:
            txtlog = open('test.txt', 'a')
            """Start"""
            root = Tk()
            root.geometry('{}x{}'.format(888, 300))
            root.title('Kanji Quiz')
            
            """top frame"""
            topFrame = Frame(root, )
            global theLabel1
            theLabel1 = Label(topFrame, text = 'Press Next!', font ='Times 100')
            theLabel2 = Label(topFrame, text = 'SHITTY QUIZ', fg = 'green')
            theLabel1.grid(row=1)
            theLabel2.grid(row=0)
            topFrame.pack()
            
            """mid frame"""
            midFrame = Frame(root)
            theLabel3 = Label(midFrame, text = 'Type \'y\' for correct or \'n\' for incorrect: ')
            entry1 = Entry(midFrame)
            theLabel3.grid(row=0)
            entry1.grid(row=1)
            midFrame.pack()
            
            """bot frame"""
            botFrame = Frame(root)
            bprev = Button(botFrame, text= 'Prev')
            bnext = Button(botFrame, text= 'Next')
            bmenu = Button(botFrame, text= 'Menu')
            bshow = Button(botFrame, text= 'Show Answer') #no parenthesis!
            bshow.bind("<Button-1>", Quiz.showAnswer) #Button-1 means a left click on the button
            bnext.bind("<Button-1>", Quiz.nextOne)
            bprev.bind("<Button-1>", Quiz.prevOne)
            bmenu.bind("<Button-1>", None)
            bprev.grid(row=0, column=1)
            bnext.grid(row=0, column=3)
            bmenu.grid(row=0, column=2)
            bshow.grid(row= 0, column=4)
            botFrame.pack()
            """End"""
            root.mainloop()
            
            txtlog.write('\n' + datetime.datetime.today().strftime('%Y-%m-%d') + ' ' \
                     + datetime.datetime.today().strftime('%T') + ' %s'% history_log)
            txtlog.close()
        except Exception:
            txtlog.write('\n' + datetime.datetime.today().strftime('%Y-%m-%d') + \
                     ' ' + datetime.datetime.today().strftime('%T') + ' An Error occured. recording... %s'% history_log)
            txtlog.close()
            

a = Quiz()
a.runtk()